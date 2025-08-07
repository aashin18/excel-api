import pandas as pd
import os
import re
import base64
from io import BytesIO
from openpyxl.styles import PatternFill
from flask import Flask, request, jsonify

app = Flask(__name__)

def load_location_data():
    try:
        if os.path.exists("lists.csv"):
            data = pd.read_csv("lists.csv")
            print(f"Successfully loaded {len(data)} locations from lists.csv")
            return data
        else:
            print("Error: lists.csv file not found.")
            return pd.DataFrame()
    except Exception as e:
        print(f"Error reading lists.csv: {str(e)}")
        return pd.DataFrame()

def check_address_match(address, dst_fac_code, location_data):
    if pd.isna(address) or str(address).strip() == '':
        return 'No Match', []

    address_to_check = str(address).strip().lower()
    dst_fac_code = str(dst_fac_code).strip().upper()

    normalized_address = address_to_check
    for char in ['-', ',', '.', '/', '\\', '&', '+', '(', ')', '[', ']', '{', '}', '|', ';', ':', '"', "'", '']:
        normalized_address = normalized_address.replace(char, ' ')
    normalized_address = ' '.join(normalized_address.split())

    matched_codes = set()
    matched_locations = []

    for _, row in location_data.iterrows():
        location = str(row['location']).strip().lower()
        code = str(row['Code']).strip().upper()

        if not location:
            continue

        normalized_location = location
        for char in ['-', ',', '.', '/', '\\', '&', '+', '(', ')', '[', ']', '{', '}', '|', ';', ':', '"', "'", '']:
            normalized_location = normalized_location.replace(char, ' ')
        normalized_location = ' '.join(normalized_location.split())

        match_found = False
        escaped_location = re.escape(normalized_location)
        pattern = r'\b' + escaped_location + r'\b'

        if re.search(pattern, normalized_address):
            match_found = True

        if match_found:
            matched_codes.add(code)
            matched_locations.append((normalized_location, code))

    if len(matched_codes) > 1:
        return 'Multi Match', matched_locations
    elif len(matched_codes) == 1:
        matched_code = list(matched_codes)[0]
        if matched_code == dst_fac_code:
            return 'Match', matched_locations
        else:
            return 'Mismatch', matched_locations
    else:
        return 'No Match', []

def process_excel_file():
    try:
        if not os.path.exists('input_file.xlsx'):
            print("Error: input_file.xlsx not found.")
            return

        location_data = load_location_data()
        if location_data.empty:
            print("Cannot proceed without location data.")
            return

        df = pd.read_excel('input_file.xlsx', engine='openpyxl')
        print(f"Input file shape: {df.shape}")

        df.iloc[0] = None
        df.iloc[1] = None

        target_columns = ["HWB", "Dst Fac", "Prod Cd", "# Pcs\\Tot Pcs", "Wt",
                          "Cnee Addr Ln 1", "Cnee Addr Ln 2", "Cnee Addr Ln 3"]

        columns_to_keep = []
        for col_index, value in enumerate(df.iloc[2]):
            value_str = str(value).strip()
            for target in target_columns:
                if value_str.lower() == target.lower():
                    columns_to_keep.append(col_index)
                    print(f"Found '{target}' in column {col_index}")
                    break

        if not columns_to_keep:
            print("No target columns found in 3rd row!")
            return

        df_filtered = df.iloc[:, columns_to_keep]

        new_headers = [str(df_filtered.iloc[2, col_idx]).strip() for col_idx in range(len(df_filtered.columns))]
        df_filtered.columns = new_headers

        addr_columns = ["Cnee Addr Ln 1", "Cnee Addr Ln 2", "Cnee Addr Ln 3"]
        existing_addr_cols = [col for col in addr_columns if col in df_filtered.columns]

        if existing_addr_cols:
            df_filtered['Address'] = df_filtered[existing_addr_cols].apply(
                lambda row: ' '.join([str(val) for val in row if pd.notna(val) and str(val).strip() != '']),
                axis=1
            )
            df_filtered = df_filtered.drop(columns=existing_addr_cols)
            print(f"Merged {len(existing_addr_cols)} address columns into 'Address'")

        df_final = df_filtered.iloc[3:].reset_index(drop=True)

        multi_match_details = []

        if 'Address' in df_final.columns and 'Dst Fac' in df_final.columns:
            validity_results = df_final.apply(
                lambda row: check_address_match(row['Address'], row['Dst Fac'], location_data),
                axis=1
            )

            df_final['Validity'] = [result[0] for result in validity_results]

            for idx, result in enumerate(validity_results):
                validity, matches = result
                if validity == 'Multi Match':
                    address = df_final.iloc[idx]['Address']
                    multi_match_details.append({
                        'row': idx + 1,
                        'address': address,
                        'matches': matches
                    })

            validity_counts = df_final['Validity'].value_counts()
            print("\nValidity Statistics:")
            print(f"- Match: {validity_counts.get('Match', 0)}")
            print(f"- Mismatch: {validity_counts.get('Mismatch', 0)}")
            print(f"- No Match: {validity_counts.get('No Match', 0)}")
            print(f"- Multi Match: {validity_counts.get('Multi Match', 0)}")

            if multi_match_details:
                print("\n=== MULTI-MATCH DETAILS ===")
                for detail in multi_match_details:
                    print(f"\nRow {detail['row']} - Address: {detail['address']}")
                    print("Matched locations and codes:")
                    for location, code in detail['matches']:
                        print(f"  - '{location}' -> Code: {code}")

        with pd.ExcelWriter('output_file.xlsx', engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

            if 'Validity' in df_final.columns:
                for row_num in range(2, len(df_final) + 2):
                    validity_value = df_final.iloc[row_num - 2]['Validity']

                    if validity_value == 'Match':
                        fill_color = green_fill
                    elif validity_value == 'Mismatch':
                        fill_color = red_fill
                    elif validity_value == 'No Match':
                        fill_color = grey_fill
                    elif validity_value == 'Multi Match':
                        fill_color = blue_fill
                    else:
                        continue

                    for col_num in range(1, len(df_final.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.fill = fill_color

        print("\nProcessing completed successfully!")
        print("Output saved to output_file.xlsx")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

@app.route('/')
def home():
    return "Flask app is running."

@app.route('/check_address', methods=['POST'])
def check_address():
    data = request.get_json()
    address = data.get('address')
    dst_fac_code = data.get('dst_fac_code')
    location_data = load_location_data()
    validity, matches = check_address_match(address, dst_fac_code, location_data)
    return jsonify({
        "validity": validity,
        "matches": matches
    })

@app.route('/process_base64', methods=['POST'])
def process_base64():
    try:
        data = request.get_json()
        input_b64 = data.get("input_file_base64")

        # Decode and save input file
        with open("input_file.xlsx", "wb") as f:
            f.write(base64.b64decode(input_b64))

        # Process it
        process_excel_file()

        # Read and encode output file
        with open("output_file.xlsx", "rb") as f:
            output_b64 = base64.b64encode(f.read()).decode("utf-8")

        return jsonify({
            "status": "success",
            "output_file_base64": output_b64
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port)
